from pathlib import Path
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / 'submission_package' / 'submission_workbook.xlsx'
OUT.parent.mkdir(exist_ok=True)

navy = '1F4E79'
light = 'D9EAF7'
warning = 'FFF2CC'
green = 'E2F0D9'
white = 'FFFFFF'
thin = Side(style='thin', color='B7C9D6')

wb = Workbook()
ws = wb.active
ws.title = 'Overview'
ws.sheet_view.showGridLines = False
ws['B2'] = 'Submission Workbook — SAC Nitrate-Reduction DFT Benchmark'
ws['B2'].font = Font(name='Georgia', size=18, bold=True, color=navy)
ws.merge_cells('B2:H2')
ws['B4'] = 'Recommended venue'
ws['C4'] = 'Computational Materials Science'
ws['B5'] = 'Article framing'
ws['C5'] = 'Reproducibility-stage periodic-DFT benchmark and production-readiness framework'
ws['B6'] = 'Repository'
ws['C6'] = 'https://github.com/hassanraza147/single-atom-nitrate-reduction-dft'
ws['B7'] = 'Latest repository commit'
ws['C7'] = '78fed93'
ws['B9'] = 'Key evidence boundary'
ws['C9'] = 'No accepted catalyst ranking, nitrate adsorption dataset, complete mechanism, selectivity, limiting potential, or electrochemical stability result is currently available.'
ws.merge_cells('C9:H10')
ws['C9'].alignment = Alignment(wrap_text=True, vertical='top')
ws['B12'] = 'Workbook contents'
for i, name in enumerate(['Submission Checklist', 'Claim Evidence', 'Run Status', 'Production Manifest', 'Convergence'], start=13):
    c = ws.cell(i, 2, name)
    c.hyperlink = f"#'{name}'!A1"
    c.font = Font(color='0563C1', underline='single')
for cell in ['B4','B5','B6','B7','B9','B12']:
    ws[cell].font = Font(bold=True, color=navy)
for col, width in {'A':3,'B':26,'C':42,'D':18,'E':18,'F':18,'G':18,'H':18}.items(): ws.column_dimensions[col].width = width


def add_csv_sheet(title, path, table_name):
    sh = wb.create_sheet(title)
    sh.sheet_view.showGridLines = False
    with path.open(newline='', encoding='utf-8') as f:
        rows = list(csv.reader(f))
    for r, row in enumerate(rows, start=1):
        for c, value in enumerate(row, start=1):
            cell = sh.cell(r, c, value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = Border(bottom=thin)
            if r == 1:
                cell.font = Font(bold=True, color=white)
                cell.fill = PatternFill('solid', fgColor=navy)
    if len(rows) > 1 and len(rows[0]) > 0:
        ref = f'A1:{get_column_letter(len(rows[0]))}{len(rows)}'
        tab = Table(displayName=table_name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True, showColumnStripes=False)
        sh.add_table(tab)
    sh.freeze_panes = 'A2'
    for c in range(1, len(rows[0])+1):
        sh.column_dimensions[get_column_letter(c)].width = min(max(14, max(len(str(rows[r][c-1])) if c-1 < len(rows[r]) else 0 for r in range(min(len(rows), 80)))), 42)
    return sh

# Submission checklist
sh = wb.create_sheet('Submission Checklist')
sh.sheet_view.showGridLines = False
headers = ['Item','Required action','Status','Author action']
for c,h in enumerate(headers,1):
    cell=sh.cell(1,c,h); cell.font=Font(bold=True,color=white); cell.fill=PatternFill('solid',fgColor=navy)
items = [
('Main manuscript','Upload Word/PDF review file with embedded components','Prepared','Confirm final author list and upload'),
('Highlights','Paste five prepared highlights','Prepared','Review wording'),
('Graphical abstract','Use workflow graphic only; do not imply unsupported performance','Specification prepared','Create/approve optional graphic'),
('Supporting Information','Upload as separate supplementary file','Prepared','Convert to portal-accepted format if required'),
('Cover letter','Address editor and name Computational Materials Science','Prepared','Insert date and corresponding-author details'),
('Research-data statement','Include repository and VASP licensing boundary','Prepared','Confirm archival DOI if required'),
('Funding statement','Provide actual funding or no-specific-funding statement','Pending author input','Complete accurately'),
('Competing interests','Provide actual declaration','Pending author input','Complete accurately'),
('Author contributions','Provide actual contribution statement','Pending author input','Complete accurately'),
('Submission declaration','Confirm originality and no concurrent submission','Pending author input','Confirm before upload'),
]
for r,row in enumerate(items,2):
    for c,v in enumerate(row,1):
        cell=sh.cell(r,c,v); cell.alignment=Alignment(wrap_text=True,vertical='top'); cell.border=Border(bottom=thin)
    if row[2]=='Prepared': sh.cell(r,3).fill=PatternFill('solid',fgColor=green)
    else: sh.cell(r,3).fill=PatternFill('solid',fgColor=warning)
sh.freeze_panes='A2'
for col,width in {'A':28,'B':58,'C':22,'D':42}.items(): sh.column_dimensions[col].width=width

add_csv_sheet('Claim Evidence', ROOT/'claim_evidence_matrix.csv', 'ClaimEvidenceTable')
add_csv_sheet('Run Status', ROOT/'data/parsed_run_status.csv', 'RunStatusTable')
add_csv_sheet('Production Manifest', ROOT/'data/production_campaign_manifest.csv', 'ProductionManifestTable')
add_csv_sheet('Convergence', ROOT/'data/convergence/convergence.csv', 'ConvergenceTable')

for sh in wb.worksheets:
    sh.sheet_properties.pageSetUpPr.fitToPage = True
    sh.page_setup.fitToWidth = 1
    sh.page_setup.fitToHeight = 0
    sh.sheet_properties.outlinePr.summaryBelow = True

wb.save(OUT)
print(OUT)
